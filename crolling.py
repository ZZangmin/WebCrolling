import requests
import pandas as pd
import os
import platform
import subprocess
import time
import http.client
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

def search_naver_products(keyword, client_id, client_secret, max_results=1000):
    url = "https://openapi.naver.com/v1/search/shop.json"
    headers = {
        "X-Naver-Client-Id": client_id,
        "X-Naver-Client-Secret": client_secret,
    }
    products = []
    display = 100  # 한 번에 최대 100개까지 가져올 수 있음
    start = 1

    while start <= max_results:
        params = {
            "query": keyword,
            "display": display,
            "start": start,
            "sort": "sim"  # 검색 정렬 방식 (유사도순)
        }

        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            data = response.json()
            items = data.get('items', [])
            if not items:
                break  # 더 이상 결과가 없으면 중단

            for item in items:
                link = item.get('link')
                # shipping_cost_value = get_shipping_cost(link)
                price = int(item.get('lprice'))
                
                # val = get_shipping_cost(link)
                # 네이버에서 배송비 조회 API 가 없음.
                # shipping_cost_value = 0
                # shipping_info = item.get('price_delivery__yw_We')
                # if shipping_info:
                #     shipping_text = shipping_info.get_text(strip=True)
                #     if "무료" in shipping_text:
                #         shipping_cost_value = 0
                #     else:
                #         try:
                #             # 배송비가 숫자로 표시된 경우 숫자만 추출
                #             shipping_cost_value = int(''.join(filter(str.isdigit, shipping_text)))
                #         except ValueError:
                #              shipping_cost_value = -1
                # totalPrice = price + shipping_cost_value
                
                # if shipping_cost_value == None:
                    # shipping_cost_value = -1
                # totalPrice = price + shipping_cost_value if shipping_cost_value > 0 else price 
                product_info = {
                    "name": item.get('title').replace('<b>', '').replace('</b>', ''),  # HTML 태그 제거
                    "price": price,
                    # "shipping": shipping_cost_value,
                    # "total" : totalPrice,
                    "mall_name": item.get('mallName', 'N/A'),
                    "link": link
                }
                products.append(product_info)

            start += display  # 다음 시작 위치로 이동
        else:
            print(f"Error: {response.status_code}, {response.text}")
            break

    return products
def get_shipping_cost(link, retry429 = 0):
    # headers = { 
    #            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36",
    #            "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    #            "Connection": "keep-alive",
    #            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8"
    #            }
     
    headers = { 'User-Agent': UserAgent().chrome }
    try:
        response = requests.get(link,headers)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            # 네이버 쇼핑 페이지의 배송비 정보를 파싱
            shipping_cost = 0
            shipping_info = soup.find('span', class_='deliveryFee')  # 예시 클래스명, 페이지에 따라 수정 필요
            if shipping_info:
                shipping_text = shipping_info.get_text(strip=True)
                if "무료" in shipping_text:
                    return 0
                else:
                    try:
                        # 배송비가 숫자로 표시된 경우 숫자만 추출
                        shipping_cost = int(''.join(filter(str.isdigit, shipping_text)))
                        return shipping_cost
                    except ValueError:
                        return -1
        elif response.status_code == 429:
            soup = BeautifulSoup(response.text, 'html.parser')
            print(f"Too Many Requests: Waiting before retrying {link}")
            shipping_info = soup.find('span', class_='deliveryFee') 
            time.sleep(0.1)  # 429 에러 시 0.1초 대기 후 재시도
            retry429 += 1
            if retry429 > 2:
                return -1
            return get_shipping_cost(link,retry429)  # 재시도
        else:
            print(f"Error fetching details from {link}: {response.status_code}")
            return -1
    except Exception as e:
        print(f"Error fetching details from {link}: {e}")
        return -1
    
def get_available_filename(base_filename):
    # 확장자 분리
    name, ext = os.path.splitext(base_filename)
    counter = 0
    new_filename = base_filename

    # 중복된 파일명이 존재하는지 확인하여 새로운 파일명 생성
    while os.path.exists(new_filename):
        counter += 1
        new_filename = f"{name}_{counter}{ext}"

    return new_filename

def open_excel_file(filename):
    # 운영체제에 따라 엑셀 파일 여는 방법을 다르게 설정
    try:
        if platform.system() == "Windows":
            os.startfile(filename)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(["open", filename])
        else:  # Linux
            subprocess.run(["xdg-open", filename])
    except Exception as e:
        print(f"Failed to open the Excel file: {e}")
        
def export_to_excel(products, base_filename="naver_products.xlsx", exclude_strings=None):
    # 사용 가능한 파일명 찾기
    filename = get_available_filename(base_filename)
    
    # 데이터프레임 생성
    df = pd.DataFrame(products)

    if exclude_strings is None:
        exclude_strings = []
        
    if exclude_strings:
        # OR 연산자를 사용하여 여러 문자열에 대해 필터링
        pattern = '|'.join(exclude_strings)
        df = df[~df['name'].str.contains(pattern, case=False, na=False)]
        
    # 중복된 링크를 기준으로 중복 제거
    df.drop_duplicates(subset='link', keep='first', inplace=True)
    
    # 엑셀 파일로 저장
    df.to_excel(filename, index=False)

    df['price'] = pd.to_numeric(df['price'], errors='coerce')
    
    # openpyxl을 사용해 하이퍼링크 추가
    wb = load_workbook(filename)
    ws = wb.active

    ws.auto_filter.ref = ws.dimensions
    # 링크를 추가할 열의 번호 (기본적으로 DataFrame 열 순서에 따라 설정)
    link_column = df.columns.get_loc("link") + 1  # 1-based index
    name_column = df.columns.get_loc("name") + 1  # 이름 열의 1-based index

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=link_column)
        cell_value = cell.value
        if cell_value:
            ws.cell(row=row, column=name_column).hyperlink = cell_value
            ws.cell(row=row, column=name_column).font = Font(color="0000FF", underline="single")

    # 각 열의 최대 너비를 설정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 컬럼 이름 가져오기 (예: A, B, C 등)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    # 파일 저장
    wb.save(filename)
    
    # 엑셀 파일 열기
    open_excel_file(filename)

if __name__ == "__main__":
    # 네이버 API 키 정보 (본인의 API 키로 대체)
    CLIENT_ID = ""
    CLIENT_SECRET = ""
    
    # 검색어 연산자 지원안함
    # keyword = "\"돼지갈비\"+\"1kg\""
    keyword = "한돈목살 양념구이 돼지갈비 1kg"
    exclude_keywords = ["양갈비", "프랜치랙", "숄더랙"]  # 제외하고 싶은 문자열을 리스트로 설정
   
    naver_products = search_naver_products(keyword, CLIENT_ID, CLIENT_SECRET)
    
    if naver_products:
        export_to_excel(naver_products, exclude_strings=exclude_keywords)
    else:
        print("No products found or API error.")